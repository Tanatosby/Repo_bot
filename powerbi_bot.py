"""
Bot para automatizar busqueda de IDs en Power BI y extraer datos.
Uso: python powerbi_bot.py
"""

import csv
import re
import time
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl

POWER_BI_URL = (
    "https://app.powerbi.com/groups/me/apps/6fa528be-040a-41c5-9a3f-ca28f00b1bbc"
    "/reports/06dc0ac3-538c-41c5-a036-8051490a55b4"
    "/bbcebd97a0e3fe7a2247?experience=power-bi"
)
INPUT_FILE    = "FR_70140526.csv"
EXCEL_GENERAL = "resultados_powerbi.xlsx"
EXCEL_ANALISIS = "analisis.xlsx"
LOG_FILE      = f"log_powerbi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

WAIT_CARGA    = 14   # segundos espera tras seleccionar ID
WAIT_NRC      = 10   # segundos espera para carga de tabla NRCs
TEST_MODE     = True
TEST_IDS      = ["72986227"]  #73333885 sobrescribe CSV en TEST_MODE; [] para usar CSV normal
BANNER_PERIODO = "202610"               # periodo a proyectar en SVAPROY


# =============================================================================
# Helpers de lectura / log
# =============================================================================

def leer_ids(path: str) -> list[str]:
    ids = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if row:
                ids.append(row[0].strip().strip('"'))
    return ids


def log(msg: str, archivo):
    print(msg)
    archivo.write(msg + "\n")
    archivo.flush()


# =============================================================================
# Navegación entre secciones del panel izquierdo
# =============================================================================

def navegar_seccion(page, titulo_parcial: str) -> bool:
    """Hace clic en un item del panel izquierdo que contenga titulo_parcial."""
    sel = f'span.itemName[title*="{titulo_parcial}"]'
    try:
        btn = page.locator(sel).first
        btn.wait_for(state="visible", timeout=8000)
        btn.click()
        time.sleep(6)
        return True
    except Exception as e:
        print(f"  No se encontro seccion '{titulo_parcial}': {e}")
        return False


# =============================================================================
# Slicer genérico (ID, Programa Académico, etc.)
# =============================================================================

JS_FIND_SLICER_BODY = """
(aria_label) => {
    // Compara normalizando Unicode (NFC) e ignorando mayúsculas/minúsculas
    const target = aria_label.normalize('NFC').toLowerCase();
    for (const el of document.querySelectorAll('.slicerBody')) {
        const label = (el.getAttribute('aria-label') || '').normalize('NFC').toLowerCase();
        if (label === target) return true;
    }
    return false;
}
"""

JS_LIST_SLICER_LABELS = """
() => [...document.querySelectorAll('.slicerBody')].map(e => e.getAttribute('aria-label'))
"""


def abrir_slicer(page, aria_label: str) -> bool:
    """Abre el slicer cuyo slicerBody tiene aria-label=aria_label."""
    # Primero verificar si ya está abierto (con normalización Unicode)
    if page.evaluate(JS_FIND_SLICER_BODY, aria_label):
        return True

    # Esperar a que los toggles aparezcan en el DOM (la página puede seguir cargando)
    try:
        page.wait_for_selector('.slicer-dropdown-menu', timeout=15000, state='visible')
    except PWTimeout:
        print(f"  Timeout esperando toggles de slicer para '{aria_label}'")
        return False

    toggles = page.locator('.slicer-dropdown-menu').all()
    for i, toggle in enumerate(toggles):
        try:
            toggle.click(timeout=2000)
            time.sleep(2.5)
            if page.evaluate(JS_FIND_SLICER_BODY, aria_label):
                print(f"  Toggle {i+1} abre slicer '{aria_label}'")
                return True
            # Debug: mostrar qué labels están abiertos
            labels = page.evaluate(JS_LIST_SLICER_LABELS)
            if labels:
                print(f"  Toggle {i+1} abrió: {labels} (buscaba: '{aria_label}')")
            page.keyboard.press("Escape")
            time.sleep(0.8)
        except Exception:
            continue
    print(f"  No se pudo abrir slicer '{aria_label}'")
    return False


def seleccionar_en_slicer(page, aria_label: str, valor: str) -> bool:
    """
    Abre el slicer indicado, busca 'valor' y hace clic en el primer resultado.
    Funciona para ID (valor exacto) y Programa Académico (valor = código, ej. PR08MEDH).
    """
    if not abrir_slicer(page, aria_label):
        return False

    # Enfocar el input del slicer correcto via JS (con normalización Unicode)
    resultado = page.evaluate(f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            if (!body) return 'no_body';
            const container = body.closest('.slicerContainer');
            if (!container) return 'no_container';
            const input = container.querySelector('input.searchInput');
            if (!input) return 'no_input';
            input.click();
            input.focus();
            return 'ok';
        }}
    """)
    if resultado != "ok":
        print(f"  JS focus slicer '{aria_label}': {resultado}")
        return False

    time.sleep(0.4)
    page.keyboard.press("Control+a")
    page.keyboard.press("Delete")
    time.sleep(0.2)
    page.keyboard.type(valor, delay=100)
    time.sleep(2.5)

    # Para ID: el título del item es exactamente el valor
    # Para Programa: el título empieza con el código (ej. "PR08MEDH: MEDICINA HUMANA")
    item = page.locator(f'.slicerItemContainer[title="{valor}"]').first
    if item.count() == 0:
        # Buscar por título que empiece con el valor (para programas)
        item = page.locator(f'.slicerItemContainer[title^="{valor}"]').first

    try:
        item.wait_for(state="visible", timeout=6000)
        item.click()
        return True
    except PWTimeout:
        print(f"  Item '{valor}' no encontrado en slicer '{aria_label}'")
        return False


def verificar_slicer_seleccionado(page, aria_label: str, valor: str) -> bool:
    """
    Abre el slicer, busca el valor y comprueba si está marcado.
    Solo hace clic si NO está marcado (evita desmarcar accidentalmente).
    """
    if not abrir_slicer(page, aria_label):
        return False

    # Enfocar el input y buscar el valor (con normalización Unicode)
    resultado = page.evaluate(f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            if (!body) return 'no_body';
            const input = body.closest('.slicerContainer')?.querySelector('input.searchInput');
            if (!input) return 'no_input';
            input.click(); input.focus();
            return 'ok';
        }}
    """)
    if resultado != "ok":
        return False

    page.keyboard.press("Control+a")
    page.keyboard.press("Delete")
    time.sleep(0.2)
    page.keyboard.type(valor, delay=100)
    time.sleep(2)

    # Buscar el item filtrado
    item = page.locator(f'.slicerItemContainer[title="{valor}"]').first
    if item.count() == 0:
        item = page.locator(f'.slicerItemContainer[title^="{valor}"]').first

    try:
        item.wait_for(state="visible", timeout=5000)
        selected = item.get_attribute("aria-selected")
        if selected == "true":
            print(f"  '{valor}' ya estaba seleccionado, no se hace clic")
            page.keyboard.press("Escape")
            return True
        # No está marcado: hacer clic para marcarlo
        item.click()
        time.sleep(0.5)
        selected = item.get_attribute("aria-selected")
        print(f"  '{valor}' seleccionado, aria-selected={selected}")
        return selected == "true"
    except PWTimeout:
        print(f"  No se encontró '{valor}' en slicer '{aria_label}'")
        return False


def seleccionar_programa_unico(page, aria_label: str, programa_code: str) -> bool:
    """
    Limpia el slicer y selecciona solo el programa indicado.
    Usa doble-clic en 'Seleccionar todo': 1er clic selecciona todo, 2do lo deselecciona.
    Así se garantiza estado limpio sin importar el estado previo.
    """
    if not abrir_slicer(page, aria_label):
        return False

    JS_CLICK_SELECT_ALL = f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            if (!body) return false;
            const container = body.closest('.slicerContainer');
            const selectAll = container?.querySelector('.slicerItemContainer[data-row-index="0"]');
            if (!selectAll) return false;
            selectAll.click();
            return true;
        }}
    """

    # Doble-clic en "Seleccionar todo" garantiza estado limpio:
    # - Si ninguno estaba marcado  → 1er clic marca todos, 2do los desmarca (limpio)
    # - Si algunos estaban marcados → 1er clic marca todos, 2do los desmarca (limpio)
    # - Si todos estaban marcados  → 1er clic los desmarca                  (limpio)
    ok1 = page.evaluate(JS_CLICK_SELECT_ALL)
    time.sleep(1.0)
    # Comprobar cuántos quedaron seleccionados para saber si el 2do clic es necesario
    n_sel = page.evaluate(f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            const container = body?.closest('.slicerContainer');
            return container?.querySelectorAll('.slicerItemContainer[aria-selected="true"]').length ?? -1;
        }}
    """)
    if n_sel != 0:
        # Aún hay seleccionados → 2do clic para desmarcar todo
        page.evaluate(JS_CLICK_SELECT_ALL)
        time.sleep(1.0)
    print(f"  Limpiar slicer '{aria_label}': {'ok' if ok1 else 'error_select_all'}")

    # Seleccionar el programa deseado
    ok = seleccionar_en_slicer(page, aria_label, programa_code)
    page.keyboard.press("Escape")
    time.sleep(1)
    print(f"  Programa '{programa_code}' {'seleccionado' if ok else 'NO seleccionado'}")
    return ok


# =============================================================================
# GENERAL: extraer tabla y código de programa
# =============================================================================

JS_EXTRAER_GENERAL = """
() => {
    const result = { nombre: '', headers: [], fila: [] };

    // Nombre: primera cadena alphabetica larga en las celdas
    for (const el of document.querySelectorAll('.cell, .pivotTableCellWrap')) {
        const t = el.textContent.trim();
        if (t.length > 5 && /^[A-ZÁÉÍÓÚÜÑ ]+$/.test(t) && !result.nombre) {
            result.nombre = t;
            break;
        }
    }

    // Encabezados vs datos usando los contenedores de Power BI
    const headerContainer = document.querySelector('.columnHeaders');
    const bodyContainer   = document.querySelector('.bodyCells');

    if (headerContainer && bodyContainer) {
        result.headers = Array.from(
            headerContainer.querySelectorAll('.cell, .pivotTableCellWrap')
        ).map(e => e.textContent.trim()).filter(Boolean);

        result.fila = Array.from(
            bodyContainer.querySelectorAll('.cell, .pivotTableCellWrap')
        ).map(e => e.textContent.trim()).filter(Boolean);
    } else {
        // Fallback ARIA roles
        result.headers = Array.from(document.querySelectorAll('[role="columnheader"]'))
            .map(e => e.textContent.trim()).filter(Boolean);
        result.fila = Array.from(document.querySelectorAll('[role="gridcell"]'))
            .map(e => e.textContent.trim()).filter(Boolean);
    }

    return result;
}
"""


def extraer_datos_general(page, student_id: str) -> dict | None:
    try:
        data = page.evaluate(JS_EXTRAER_GENERAL)
        print(f"  Nombre  : {data.get('nombre', '-')}")
        print(f"  Headers : {data['headers']}")
        print(f"  Valores : {data['fila']}")
        if not data["headers"]:
            print("  ADVERTENCIA: tabla GENERAL vacia")
            return None
        return data
    except Exception as e:
        print(f"  Error extrayendo GENERAL: {e}")
        return None


def obtener_programa_code(data: dict) -> str | None:
    """Extrae el código de programa (ej. 'PR08MEDH') del dict de GENERAL."""
    try:
        idx = next(
            i for i, h in enumerate(data["headers"])
            if "PROGRAMA" in h.upper() and "DESCRIPCI" not in h.upper()
        )
        return data["fila"][idx]
    except (StopIteration, IndexError):
        return None


# =============================================================================
# DETALLE DEL ÁREA: extraer cursos con celda naranja (N)
# =============================================================================

JS_CURSOS_N = """
() => {
    const cursosN = [];
    const debugInfo = [];

    // Buscar todas las celdas de la columna CMP_REGLA_MOSTRAR (columna "CUMP.")
    const cumpCells = document.querySelectorAll(
        '[data-query-ref="T03_DETALLE_AREA_INDIV.CMP_REGLA_MOSTRAR"]'
    );

    for (const cell of cumpCells) {
        const wrap = cell.querySelector('.pivotTableCellWrap');
        const texto = wrap ? wrap.textContent.trim() : cell.textContent.trim();
        debugInfo.push({ texto: texto.slice(0, 20), ref: 'CMP_REGLA_MOSTRAR' });

        // Solo nos interesan las celdas con valor "N"
        if (texto !== 'N') continue;

        // La descripcion del curso esta en la misma fila [role="row"]
        const row = cell.closest('[role="row"]');
        if (!row) continue;

        const descCell = row.querySelector(
            '[data-query-ref="DETALLE_AREA.DESCRIPCION_REGLA_DA"]'
        );
        if (descCell) {
            const nombre = descCell.textContent.trim();
            // Filtrar vacios y &nbsp;
            if (nombre && nombre !== '\\u00a0' && !cursosN.includes(nombre)) {
                cursosN.push(nombre);
            }
        }
    }

    return { cursosN, debug: debugInfo.slice(0, 10), total: cumpCells.length };
}
"""


def extraer_cursos_N(page) -> list[str]:
    try:
        result = page.evaluate(JS_CURSOS_N)
        debug = result.get("debug", [])
        cursos = result.get("cursosN", [])
        print(f"  Debug celdas N detectadas ({len(debug)}): {debug}")
        print(f"  Cursos con N ({len(cursos)}): {cursos}")
        return cursos
    except Exception as e:
        print(f"  Error extrayendo cursos N: {e}")
        return []


JS_FIND_SCROLL_CONTAINER = """
() => {
    // Power BI virtualiza filas; buscar el contenedor que tiene scroll vertical
    const candidatos = [
        '.bodyCells',
        '.scrollRegion',
        '.tableEx .innerContainer',
        '.matrixScrollContainer',
        '[class*="scrollable"]',
    ];
    for (const sel of candidatos) {
        const el = document.querySelector(sel);
        // Excluir si está dentro de un popup de slicer
        if (el && !el.closest('.slicer-dropdown-popup') && el.scrollHeight > el.clientHeight + 5) {
            return { selector: sel, scrollHeight: el.scrollHeight, clientHeight: el.clientHeight };
        }
    }
    // Fallback: buscar cualquier div con overflow scrolleable (excluir slicers)
    for (const el of document.querySelectorAll('div')) {
        if (el.closest('.slicer-dropdown-popup')) continue;
        const style = window.getComputedStyle(el);
        if ((style.overflowY === 'auto' || style.overflowY === 'scroll')
            && el.scrollHeight > el.clientHeight + 20
            && el.clientHeight > 50) {
            el.setAttribute('data-scroll-target', 'true');
            return { selector: '[data-scroll-target="true"]', scrollHeight: el.scrollHeight, clientHeight: el.clientHeight };
        }
    }
    return null;
}
"""

JS_SCROLL_CONTAINER_TO = """
(args) => {
    const el = document.querySelector(args.selector);
    if (!el) return false;
    el.scrollTop = args.position;
    return el.scrollTop;
}
"""

JS_EXTRACT_VISIBLE_N = """
() => {
    const cursos = [];
    const cells = document.querySelectorAll(
        '[data-query-ref="T03_DETALLE_AREA_INDIV.CMP_REGLA_MOSTRAR"]'
    );
    for (const cell of cells) {
        const wrap = cell.querySelector('.pivotTableCellWrap');
        const texto = wrap ? wrap.textContent.trim() : cell.textContent.trim();
        if (texto !== 'N') continue;

        const row = cell.closest('[role="row"]');
        if (!row) continue;

        const descCell = row.querySelector('[data-query-ref="DETALLE_AREA.DESCRIPCION_REGLA_DA"]');
        if (descCell) {
            const nombre = descCell.textContent.trim();
            if (nombre && nombre !== ' ') cursos.push(nombre);
        }
    }
    return cursos;
}
"""


def extraer_cursos_N_con_scroll(page) -> list[str]:
    """Scrollea la tabla de DETALLE DEL ÁREA completa para encontrar todos los cursos con N."""
    todos = set()

    info = page.evaluate(JS_FIND_SCROLL_CONTAINER)
    if not info:
        print("  No se encontró contenedor scrolleable, extrayendo solo vista actual")
        return extraer_cursos_N(page)

    sel = info["selector"]
    scroll_height = info["scrollHeight"]
    client_height = info["clientHeight"]
    step = max(client_height - 40, 80)   # solapa ~40px para no perder filas del borde
    total_steps = max(1, int(scroll_height / step) + 2)
    print(f"  Scroll container: '{sel}' | scrollHeight={scroll_height} | paso={step} | ~{total_steps} pasos")

    position = 0
    page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": 0})
    time.sleep(0.6)

    while True:
        visibles = page.evaluate(JS_EXTRACT_VISIBLE_N)
        antes = len(todos)
        todos.update(visibles)
        if len(todos) > antes:
            print(f"    pos={position}: +{len(todos)-antes} nuevos -> total {len(todos)}")

        # ¿Llegamos al final?
        if position + client_height >= scroll_height:
            break

        position = min(position + step, scroll_height - client_height)
        page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": position})
        time.sleep(0.5)

    # Devolver scroll al inicio
    page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": 0})

    cursos = sorted(todos)
    print(f"  Cursos con N tras scroll completo ({len(cursos)}): {cursos}")
    return cursos


# =============================================================================
# CURSOS PRINCIPALES: mapa completo curso → prerrequisitos
# =============================================================================

JS_EXTRACT_PREREQS_VISIBLE = """
() => {
    const result = {};
    const codigos = {};
    const cursoCells = document.querySelectorAll('[data-query-ref="Main_Table.Curso"]');
    for (const cell of cursoCells) {
        const wrap = cell.querySelector('.pivotTableCellWrap');
        const raw = (wrap || cell).textContent.trim();
        const sep = raw.indexOf(': ');
        const codigo = sep !== -1 ? raw.slice(0, sep) : '';
        const nombre = sep !== -1 ? raw.slice(sep + 2) : raw;
        if (codigo) codigos[codigo] = nombre;
        if (!nombre || nombre === ' ') continue;

        const row = cell.closest('[role="row"]');
        if (!row) continue;

        const prereqCell = row.querySelector('[data-query-ref="Main_Table.Pre-requisitos"]');
        const prereq = prereqCell
            ? (prereqCell.querySelector('.pivotTableCellWrap') || prereqCell).textContent.trim()
            : '';

        if (!(nombre in result)) {
            result[nombre] = prereq;
        }
    }
    return { result, codigos };
}
"""


def extraer_mapa_prereqs(page) -> tuple[dict, dict]:
    """Scrollea CURSOS PRINCIPALES y construye ({curso: prereqs}, {codigo: nombre})."""
    mapa = {}
    codigos = {}

    info = page.evaluate(JS_FIND_SCROLL_CONTAINER)
    if not info:
        print("  Sin scroll vertical, extrayendo vista actual")
        data = page.evaluate(JS_EXTRACT_PREREQS_VISIBLE)
        return data["result"], data["codigos"]

    sel = info["selector"]
    scroll_height = info["scrollHeight"]
    client_height = info["clientHeight"]
    step = max(client_height - 40, 80)
    print(f"  Scroll prereqs: scrollHeight={scroll_height} | paso={step}")

    page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": 0})
    time.sleep(0.6)
    position = 0

    while True:
        data = page.evaluate(JS_EXTRACT_PREREQS_VISIBLE)
        parcial = data["result"]
        codigos.update(data["codigos"])
        nuevos = {k: v for k, v in parcial.items() if k not in mapa}
        mapa.update(nuevos)
        if nuevos:
            print(f"    pos={position}: +{len(nuevos)} cursos -> total {len(mapa)}")

        if position + client_height >= scroll_height:
            break

        position = min(position + step, scroll_height - client_height)
        page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": position})
        time.sleep(0.5)

    page.evaluate(JS_SCROLL_CONTAINER_TO, {"selector": sel, "position": 0})
    print(f"  Mapa completo: {len(mapa)} cursos, {len(codigos)} codigos indexados")
    return mapa, codigos


# =============================================================================
# Lógica de proyección
# =============================================================================

# Cursos que no se proyectan directamente (se gestionan por separado)
_EXCLUIR_PREFIJOS = (
    'ELECTIVO',
    'MINOR',
    'CURSO DE LIBRE CONFIGURACI',  # cubre con/sin tilde: CONFIGURACION / CONFIGURACIÓN
    'LIBRE CONFIGURACI',
)

def es_curso_excluido(nombre: str) -> bool:
    n = nombre.strip().upper()
    return any(n.startswith(p) for p in _EXCLUIR_PREFIJOS)


def calcular_proyectar(prereq_str: str, mapa_codigos: dict, cursos_n_set: set) -> str:
    """
    Evalúa si un curso puede proyectarse según sus prerrequisitos.
    prereq_str: "BIET08001 [11] AND TEOL08002 [11]" o con OR
    mapa_codigos: {codigo: nombre_curso}
    cursos_n_set: set de nombres de cursos con N (en mayúsculas)
    Devuelve "Para proyectar" o "No proyectar".
    """
    if not prereq_str or not prereq_str.strip():
        return 'Para proyectar'

    def reemplazar(match):
        codigo = match.group(1)
        nombre = mapa_codigos.get(codigo, '').upper()
        # Si no está en codigos o no está en cursos_n → cumple (True)
        return 'False' if nombre in cursos_n_set else 'True'

    # Reemplazar "CODIGO [N]" o solo "CODIGO" por True/False
    expr = re.sub(r'([A-Z]{2,5}\d{5})(?:\s*\[\d+\])?', reemplazar, prereq_str)
    expr = expr.replace(' AND ', ' and ').replace(' OR ', ' or ')

    try:
        cumple = eval(expr)  # expr solo tiene True/False/and/or/()
        return 'Para proyectar' if cumple else 'No proyectar'
    except Exception:
        return 'Revisar'


# =============================================================================
# Excel helpers
# =============================================================================

def guardar_general(student_id: str, data: dict):
    path = Path(EXCEL_GENERAL)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(["ID", "NOMBRE"] + data["headers"])
        print(f"  Excel GENERAL creado con headers: {['ID','NOMBRE']+data['headers']}")

    ws.append([student_id, data.get("nombre", "")] + data["fila"])
    wb.save(path)
    print(f"  Guardado en {EXCEL_GENERAL}")


def guardar_analisis(filas: list[dict]):
    """Agrega filas a analisis.xlsx. Cada dict: {id, codigo, curso, prerequisitos, proyectar}"""
    if not filas:
        return
    path = Path(EXCEL_ANALISIS)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analisis"
        ws.append(["ID", "Código", "Curso no cumple", "Prerrequisitos", "Proyectar"])
        print(f"  Excel ANALISIS creado")

    for f in filas:
        ws.append([f["id"], f.get("codigo", ""), f["curso"], f["prerequisitos"], f.get("proyectar", "")])
    try:
        wb.save(path)
        print(f"  {len(filas)} fila(s) guardadas en {EXCEL_ANALISIS}")
    except PermissionError:
        alt = path.with_stem(path.stem + "_" + datetime.now().strftime("%H%M%S"))
        wb.save(alt)
        print(f"  ADVERTENCIA: {EXCEL_ANALISIS} estaba abierto, guardado en {alt.name}")


EXCEL_SVAPROY = "analisis_svaproy.xlsx"


def guardar_analisis_svaproy(student_id: str,
                              cursos_bot: list[dict],
                              filas_svaproy: list[dict]):
    """
    Cruza cursos 'Para proyectar' del bot con los que ya están en SVAPROY.
    cursos_bot: lista de dicts {nombre, codigo}  (codigo = ej. 'INVE01010')
    Columnas: ID | Código | Nombre (bot) | En SVAPROY | Descripción SVAPROY
    """
    def _norm(s: str) -> str:
        return s.replace(' ', '').upper().strip()

    # Indexar SVAPROY por Materia+Curso normalizado → fila completa
    svaproy_idx = {}
    for r in filas_svaproy:
        materia = (r.get('Materia') or '').strip()
        curso   = (r.get('Curso') or '').strip()
        key = _norm(materia + curso)
        if key:
            svaproy_idx[key] = r

    print(f"  SVAPROY códigos indexados ({len(svaproy_idx)}): {sorted(svaproy_idx.keys())[:8]}")

    path = Path(EXCEL_SVAPROY)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparacion"
        ws.append(["ID", "Código", "Nombre (bot)", "En SVAPROY", "Descripción SVAPROY"])

    coinciden = 0
    sin_svaproy = []
    for item in cursos_bot:
        codigo = item.get("codigo", "")
        nombre = item.get("nombre", "")
        en_svaproy = "No"
        desc_orig  = ""

        fila_match = svaproy_idx.get(_norm(codigo)) if codigo else None

        if fila_match:
            en_svaproy = "Sí"
            coinciden += 1
            desc_col = next((k for k in fila_match if "escripci" in k.lower()), None)
            desc_orig = (
                fila_match.get(desc_col, '').strip() if desc_col
                else next((v for v in fila_match.values() if v and v.strip()), "")
            )
        else:
            sin_svaproy.append(item)

        ws.append([student_id, codigo, nombre, en_svaproy, desc_orig])

    try:
        wb.save(path)
        total = len(cursos_bot)
        print(f"  analisis_svaproy.xlsx: {total} cursos bot | "
              f"{coinciden} en SVAPROY | {total - coinciden} no en SVAPROY")
    except PermissionError:
        alt = path.with_stem(path.stem + "_" + datetime.now().strftime("%H%M%S"))
        wb.save(alt)
        print(f"  ADVERTENCIA: archivo abierto, guardado como {alt.name}")
    return sin_svaproy


# =============================================================================
# Banner: navegación a SVAPROY
# =============================================================================

BANNER_URL = "https://banner.udep.edu.pe/applicationNavigator/seamless"

JS_LEER_SVAPROY = """
() => {
    const result = { headers: [], rows: [], debug: [] };

    // ─── Diagnóstico: todos los inputs con valor ───────────────────────────
    const allInputs = [...document.querySelectorAll('input')];
    result.debug.push('inputs total: ' + allInputs.length);
    for (const inp of allInputs) {
        const v = (inp.value || inp.getAttribute('value') || '').trim();
        if (v) result.debug.push(
            'inp[' + (inp.name || inp.id || '?').slice(0,30) + ']="' + v.slice(0,50) + '"' +
            (inp.readOnly ? ' RO' : ''));
    }

    // ─── Estrategia A: tabla con datos (excluir la de paginación "Ir a…") ──
    const allTables = [...document.querySelectorAll('table')];
    result.debug.push('tablas: ' + allTables.length);
    let dataTable = null, bestScore = -1;
    for (let i = 0; i < allTables.length; i++) {
        const t = allTables[i];
        const txt = t.textContent.toLowerCase();
        // Saltar la tabla de paginación (solo tiene "ir a")
        if (txt.includes('ir a') && t.querySelectorAll('tr').length <= 4
            && t.querySelectorAll('input[type="text"]').length <= 1) continue;
        const inputCount = t.querySelectorAll(
            'input:not([type="button"]):not([type="submit"])' +
            ':not([type="image"]):not([type="checkbox"]):not([type="hidden"])').length;
        const rowCount = t.querySelectorAll('tr').length;
        const score = inputCount * 10 + rowCount;
        result.debug.push('t['+i+']: rows='+rowCount+' inputs='+inputCount);
        if (score > bestScore && rowCount > 0) { bestScore = score; dataTable = t; }
    }

    if (dataTable && bestScore > 0) {
        const thEls = [...dataTable.querySelectorAll('th')];
        if (thEls.length > 0) {
            result.headers = thEls.map(th => th.textContent.trim()).filter(Boolean);
        } else {
            const firstRow = dataTable.querySelector('tr');
            if (firstRow)
                result.headers = [...firstRow.querySelectorAll('td')]
                    .map(td => td.textContent.trim()).filter(Boolean);
        }
        let bodyRows = [...dataTable.querySelectorAll('tbody tr')];
        if (bodyRows.length === 0)
            bodyRows = [...dataTable.querySelectorAll('tr')].slice(result.headers.length > 0 ? 1 : 0);
        for (const row of bodyRows) {
            const cells = [...row.querySelectorAll('td')].map(td => {
                const inp = td.querySelector(
                    'input:not([type="button"]):not([type="submit"])' +
                    ':not([type="image"]):not([type="checkbox"]):not([type="hidden"])');
                if (inp) {
                    return (inp.value || inp.getAttribute('value') || td.textContent).trim();
                }
                return td.textContent.trim();
            });
            if (cells.some(c => c && c.trim())) result.rows.push(cells);
        }
        if (result.rows.length > 0) {
            result.debug.push('A: filas=' + result.rows.length);
            return result;
        }
    }

    // ─── Estrategia B: inputs fuera de tabla (div-based Banner data block) ─
    // Banner 9 INB a veces renderiza el bloque de datos como divs con inputs.
    // Agrupamos inputs por su ancestro contenedor de "fila" (div/section con
    // varios inputs hermanos), usando el nombre del input para detectar registros.
    result.debug.push('Estrategia B: inputs fuera de tabla');

    // Recolectar inputs que NO estén dentro de ninguna tabla
    const inputsFueraDeTbl = allInputs.filter(inp => !inp.closest('table'));
    result.debug.push('inputs fuera de tabla: ' + inputsFueraDeTbl.length);

    if (inputsFueraDeTbl.length > 0) {
        // Agrupar por ancestro que contenga ≥2 inputs hermanos (= una fila del data block)
        const seen = new Set();
        const grupos = [];
        for (const inp of inputsFueraDeTbl) {
            let ancestor = inp.parentElement;
            for (let d = 0; d < 8 && ancestor; d++) {
                const cnt = ancestor.querySelectorAll('input').length;
                if (cnt >= 2) {
                    if (!seen.has(ancestor)) {
                        seen.add(ancestor);
                        grupos.push(ancestor);
                    }
                    break;
                }
                ancestor = ancestor.parentElement;
            }
        }
        result.debug.push('grupos B: ' + grupos.length);
        for (const g of grupos) {
            const vals = [...g.querySelectorAll('input')].map(
                i => (i.value || i.getAttribute('value') || '').trim());
            if (vals.some(v => v)) result.rows.push(vals);
        }
        if (result.rows.length > 0) {
            result.debug.push('B: filas=' + result.rows.length);
            return result;
        }
    }

    // ─── Estrategia C: dump plano de todos los inputs con valor ───────────
    result.debug.push('Estrategia C: dump plano');
    const vals = allInputs
        .map(i => (i.value || i.getAttribute('value') || '').trim())
        .filter(v => v);
    if (vals.length > 0) result.rows.push(vals);
    result.debug.push('C: vals=' + vals.length);

    return result;
}
"""


def leer_datos_svaproy(banner_page) -> list[dict]:
    """
    Lee la tabla de proyección de SVAPROY tras hacer clic en Ir.
    Devuelve lista de dicts {columna: valor} por cada fila.
    """
    form = _svaproy_frame(banner_page)
    if not form:
        print("  No se encontró frame para leer SVAPROY")
        return []
    try:
        # Dar tiempo extra a Banner para terminar de poblar los inputs
        time.sleep(2)
        # Banner INB puede no tener <tbody>; esperar cualquier <tr> dentro de <table>
        try:
            form.wait_for_selector('table tr', timeout=8000, state='attached')
        except Exception:
            print("  Advertencia: no aparecieron filas de tabla, continuando de todos modos")

        data    = form.evaluate(JS_LEER_SVAPROY)
        debug   = data.get('debug', [])
        headers = data.get('headers', [])
        rows    = data.get('rows', [])

        print(f"  SVAPROY debug: {debug}")
        print(f"  SVAPROY columnas ({len(headers)}): {headers}")
        filas = []
        for row in rows:
            fila = {(headers[i] if i < len(headers) else f"col{i}"): v
                    for i, v in enumerate(row)}
            print(f"    {fila}")
            filas.append(fila)
        print(f"  Total filas en SVAPROY: {len(filas)}")
        return filas
    except Exception as e:
        print(f"  Error leyendo datos SVAPROY: {e}")
        return []


def _parsear_export_svaproy(path: Path) -> list[dict]:
    """
    Parsea el archivo XLSX/CSV descargado de Banner SVAPROY con SHIFT+F1.

    Estructura Banner:
      - Fila 1: bloque clave (ID alumno, nombre, periodo, plan) — NO son encabezados
      - Fila 2 (opcional): encabezados de columna si Banner los incluye
      - Filas restantes: una por curso proyectado

    Columnas conocidas de SVAPROY (por posición):
      0=Plan de estudios  1=Programa  2=Materia  3=Curso  4=Descripción
      5=Nivel  6=Atributo  7=Prioridad de área
    """
    import csv as csv_mod

    SVAPROY_COLS = [
        'Plan de estudios', 'Programa', 'Materia', 'Curso',
        'Descripción', 'Nivel', 'Atributo', 'Prioridad de área',
    ]
    filas = []

    def _row_to_dict(row: list) -> dict:
        """Construye dict con nombres de columna SVAPROY; evita colisión de claves vacías."""
        d = {}
        for i, v in enumerate(row):
            key = SVAPROY_COLS[i] if i < len(SVAPROY_COLS) else f'col{i}'
            d[key] = v
        return d

    def _parece_header(row: list) -> bool:
        """True si la fila parece ser una fila de encabezados (contiene 'Plan', 'Programa', etc.)"""
        joined = ' '.join(str(c or '') for c in row).lower()
        return any(kw in joined for kw in ('plan', 'programa', 'materia', 'descripci'))

    def _procesar_filas_raw(raw: list[list]) -> list[dict]:
        if not raw:
            return []
        # Saltar siempre la primera fila (bloque clave del alumno)
        rest = raw[1:]
        # Si la primera de las restantes parece encabezados de Banner, saltarla también
        if rest and _parece_header(rest[0]):
            rest = rest[1:]
        result = []
        for row in rest:
            if not any(v for v in row):
                continue   # fila vacía
            result.append(_row_to_dict(row))
        return result

    # ── Intento 1: Excel ──────────────────────────────────────────────────
    try:
        wb   = openpyxl.load_workbook(path, data_only=True)
        ws   = wb.active
        raw  = [[str(c or '').strip() for c in row] for row in ws.iter_rows(values_only=True)]
        filas = _procesar_filas_raw(raw)
        if filas:
            print(f"  Formato Excel | {len(filas)} filas")
            for f in filas[:6]:
                print(f"    {f}")
            return filas
    except Exception as e_xls:
        print(f"  No es Excel ({e_xls}), intentando CSV...")

    # ── Intento 2: CSV ────────────────────────────────────────────────────
    for enc in ('utf-8-sig', 'latin-1', 'utf-8'):
        try:
            with open(path, newline='', encoding=enc) as f:
                raw = [[v.strip() for v in row] for row in csv_mod.reader(f)]
            filas = _procesar_filas_raw(raw)
            if filas:
                print(f"  Formato CSV ({enc}) | {len(filas)} filas")
                for f in filas[:6]:
                    print(f"    {f}")
                break
        except (UnicodeDecodeError, Exception) as e_csv:
            filas = []

    print(f"  Export parseado: {len(filas)} filas")
    return filas


def exportar_svaproy(banner_page) -> list[dict]:
    """
    Exporta datos de SVAPROY con SHIFT+F1 (Banner INB extract).
    Prioridades: (1) descarga de archivo, (2) nueva pestaña/popup, (3) lectura DOM.
    """
    form = _svaproy_frame(banner_page)
    if not form:
        print("  Frame no encontrado, usando lectura DOM")
        return leer_datos_svaproy(banner_page)

    time.sleep(1)   # dejar que el data block tenga foco tras el Ir
    print("  Disparando SHIFT+F1 para exportar...")

    # ── Intento 1: descarga de archivo ────────────────────────────────────────
    try:
        with banner_page.expect_download(timeout=10000) as dl:
            banner_page.keyboard.press("Shift+F1")
        download = dl.value
        fname     = download.suggested_filename or "svaproy_export"
        ext       = Path(fname).suffix or ".xlsx"   # Banner envía UUID sin extensión
        save_path = Path(f"svaproy_export_tmp{ext}")
        download.save_as(save_path)
        print(f"  Descargado: {save_path} ({save_path.stat().st_size} bytes)")
        return _parsear_export_svaproy(save_path)
    except Exception as e1:
        print(f"  Sin descarga: {e1}")

    # ── Intento 2: Banner puede abrir popup/nueva pestaña con tabla HTML ──────
    try:
        with banner_page.context.expect_page(timeout=6000) as pg_info:
            banner_page.keyboard.press("Shift+F1")
        new_page = pg_info.value
        new_page.wait_for_load_state("networkidle", timeout=8000)
        data  = new_page.evaluate(JS_LEER_SVAPROY)
        new_page.close()
        headers = data.get('headers', [])
        rows    = data.get('rows', [])
        print(f"  Popup: {len(rows)} filas | debug: {data.get('debug', [])}")
        filas = [{(headers[i] if i < len(headers) else f"col{i}"): v
                  for i, v in enumerate(row)} for row in rows]
        for fila in filas[:6]:
            print(f"    {fila}")
        return filas
    except Exception as e2:
        print(f"  Sin popup: {e2}")

    # ── Fallback: lectura DOM ─────────────────────────────────────────────────
    print("  Fallback: lectura DOM")
    return leer_datos_svaproy(banner_page)


def _svaproy_frame(banner_page):
    """
    Espera hasta 15 s a que aparezca el iframe del formulario Banner INB (SVAPROY).
    Devuelve el Frame object, o None si no aparece.
    """
    for _ in range(30):
        for fr in banner_page.frames:
            if fr.url == banner_page.url:
                continue
            try:
                if fr.locator('input[type="text"]').count() >= 2:
                    print(f"  Frame SVAPROY detectado: {fr.url[:80]}")
                    return fr
            except Exception:
                pass
        time.sleep(0.5)
    return None


def navegar_svaproy(banner_page) -> bool:
    """Abre el buscador lateral de Banner, escribe SVAPROY y entra a la página."""
    try:
        # 1. Clic en el botón "Buscar" del sidebar
        banner_page.locator("#sidebarSearchLink").click()
        banner_page.wait_for_selector("#searchMenu", state="visible", timeout=8000)
        time.sleep(0.8)

        # 2. Clic en el input de búsqueda y escribir SVAPROY
        search_input = banner_page.locator("input#search")
        search_input.wait_for(state="visible", timeout=5000)
        search_input.click()
        time.sleep(0.3)
        search_input.fill("SVAPROY")
        time.sleep(2)  # esperar que aparezcan resultados

        # 3. Clic en el primer resultado que coincida con SVAPROY
        resultado = banner_page.locator("#vsearchResultId li").first
        resultado.wait_for(state="visible", timeout=8000)
        print(f"  Banner resultado: {resultado.inner_text()[:80]}")
        resultado.click()

        # 4. Esperar que cargue la página SVAPROY dentro del iframe o marco
        time.sleep(3)
        print(f"  Banner URL tras SVAPROY: {banner_page.url}")
        return True
    except Exception as e:
        print(f"  Error navegando a SVAPROY: {e}")
        return False


def rellenar_svaproy(banner_page, student_id: str, periodo: str) -> bool:
    """
    Rellena el formulario SVAPROY para un alumno:
      ID → Tab (resuelve nombre) → Periodo → Tab → "..." Plan de estudios
      → selecciona primer plan → OK → Ir
    El formulario Banner INB vive en un iframe dentro de applicationNavigator.
    """
    try:
        form = _svaproy_frame(banner_page)
        if not form:
            print("  Frame no encontrado, usando página principal")
            form = banner_page

        # ── 1. Campo ID ──────────────────────────────────────────────────────
        id_input = form.locator('input[type="text"]').first
        id_input.wait_for(state="visible", timeout=8000)
        id_input.click()
        id_input.press("Control+a")
        id_input.fill(student_id)
        id_input.press("Tab")
        time.sleep(2.5)                  # espera resolución del nombre

        # ── 2. Campo Periodo ──────────────────────────────────────────────────
        # Después de Tab desde ID, el foco de Banner queda en Periodo.
        # page.keyboard envía al elemento activo del navegador (funciona cross-frame).
        banner_page.keyboard.press("Control+a")
        banner_page.keyboard.type(periodo, delay=60)
        banner_page.keyboard.press("Tab")
        time.sleep(1)

        # ── 3. Clic en campo Plan de estudios y abrir lookup con F9 ──────────
        # Tab desde Periodo salta Plan de estudios (va al botón Ir).
        # Hay que hacer clic directo en el campo para activarlo, luego F9.
        print("  Activando campo Plan de estudios...")
        click_result = form.evaluate("""
            () => {
                // Buscar por etiqueta label/td/span con texto "Plan de estudios"
                for (const el of document.querySelectorAll('label, td, span, div, th')) {
                    if (/^Plan de estudios/i.test(el.textContent.trim()) &&
                            el.children.length === 0) {
                        const contenedor = el.closest('tr') || el.parentElement;
                        if (contenedor) {
                            const inp = contenedor.querySelector(
                                'input[type="text"], input:not([type="button"]):not([type="submit"]):not([type="image"])');
                            if (inp && !inp.readOnly) {
                                inp.click(); inp.focus(); return 'ok:label';
                            }
                        }
                    }
                }
                // Fallback: último input editable del formulario
                const inputs = [...document.querySelectorAll(
                    'input[type="text"], input:not([type="button"]):not([type="submit"]):not([type="image"])')
                ].filter(i => !i.readOnly);
                const last = inputs[inputs.length - 1];
                if (last) { last.click(); last.focus(); return 'fallback:last'; }
                return 'not_found';
            }
        """)
        print(f"  Plan de estudios activado: {click_result}")
        time.sleep(0.4)
        print("  Abriendo lookup con F9...")
        banner_page.keyboard.press("F9")
        time.sleep(2.5)

        # ── 4. Modal "Plan de estudios (SORLCUR)" ────────────────────────────
        modal_ctx = None
        for ctx in [form, banner_page]:
            try:
                ctx.wait_for_selector('text=SORLCUR', timeout=6000, state='visible')
                modal_ctx = ctx
                break
            except PWTimeout:
                continue

        if not modal_ctx:
            print("  Modal Plan de estudios no apareció")
            return False

        print("  Modal encontrado, seleccionando primera fila...")
        # Usar JS para clic: evita el check de visibilidad de Playwright
        # (Banner INB puede tener filas "ocultas" para criterios de filtro antes del primer plan)
        fila_result = modal_ctx.evaluate("""
            () => {
                // Primera fila de tbody con altura > 0 (visible)
                for (const row of document.querySelectorAll('tbody tr')) {
                    const rect = row.getBoundingClientRect();
                    if (rect.height > 0) {
                        row.click();
                        return 'ok:' + row.textContent.trim().slice(0, 60);
                    }
                }
                // Fallback: segunda tr (saltar header)
                const trs = document.querySelectorAll('tr');
                if (trs.length > 1) {
                    trs[1].click();
                    return 'fallback:' + trs[1].textContent.trim().slice(0, 60);
                }
                return 'not_found';
            }
        """)
        print(f"  Fila seleccionada: {fila_result}")
        time.sleep(0.5)

        # Clic en OK — Banner INB usa <input value="OK"> o <button>
        print("  Haciendo clic en OK...")
        ok_clicked = False
        for sel in ['input[value="OK"]', 'button:has-text("OK")', 'input[value="Ok"]']:
            try:
                btn = modal_ctx.locator(sel).first
                if btn.count() > 0:
                    btn.click()
                    ok_clicked = True
                    print(f"  OK con: {sel}")
                    break
            except Exception:
                continue

        if not ok_clicked:
            # Debug: mostrar todos los botones del modal para diagnóstico
            btns = modal_ctx.evaluate("""
                () => [...document.querySelectorAll('button, input[type="button"], input[type="submit"]')]
                    .map(el => ({tag: el.tagName, text: el.textContent.trim(),
                                 value: el.value || '', cls: el.className.substring(0,30)}))
            """)
            print(f"  Botones en modal: {btns}")
            return False
        time.sleep(1.5)

        # ── 5. Botón Ir ───────────────────────────────────────────────────────
        form.locator('button:has-text("Ir")').click()
        time.sleep(6)    # Banner INB necesita tiempo para renderizar el data block

        print(f"  SVAPROY Ir ejecutado: ID={student_id} | Periodo={periodo}")
        return True

    except Exception as e:
        print(f"  Error rellenando SVAPROY: {e}")
        return False


# =============================================================================
# Limpieza de slicer con botón "Borrar selecciones" (◇)
# =============================================================================

JS_LIMPIAR_SLICER = """
(aria_label) => {
    const target = aria_label.normalize('NFC').toLowerCase();
    const body = [...document.querySelectorAll('.slicerBody')].find(
        e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
    );
    if (!body) return 'no_body';

    // Subir por el DOM hasta 12 niveles buscando el botón "Borrar selecciones"
    let el = body.parentElement;
    for (let i = 0; i < 12; i++) {
        if (!el) break;
        const btn = el.querySelector(
            '[title="Borrar selecciones"], [aria-label="Borrar selecciones"], ' +
            '.slicerClearButton, button.clear'
        );
        if (btn) { btn.click(); return 'ok'; }
        el = el.parentElement;
    }
    return 'not_found';
}
"""

# JS para hacer hover sobre el visual container del slicer (fuerza aparición del botón)
JS_HOVER_SLICER_VISUAL = """
(aria_label) => {
    const target = aria_label.normalize('NFC').toLowerCase();
    const body = [...document.querySelectorAll('.slicerBody')].find(
        e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
    );
    if (!body) return false;
    // Subir hasta el visual container y disparar mouseover
    let el = body;
    for (let i = 0; i < 10; i++) {
        el = el.parentElement;
        if (!el) break;
        if (el.classList.contains('visual-container') ||
            el.classList.contains('visualContainerHost') ||
            el.getAttribute('data-testid')) {
            el.dispatchEvent(new MouseEvent('mouseover', {bubbles: true}));
            el.dispatchEvent(new MouseEvent('mouseenter', {bubbles: true}));
            return true;
        }
    }
    return false;
}
"""


def limpiar_slicer_borrador(page, aria_label: str) -> bool:
    """
    Limpia el slicer usando el botón 'Borrar selecciones' (◇).
    Hace hover para forzar la aparición del botón antes de buscarlo.
    Fallback: doble-clic en 'Seleccionar todo'.
    """
    # Hover para que Power BI muestre los controles del visual
    page.evaluate(JS_HOVER_SLICER_VISUAL, aria_label)
    time.sleep(0.5)

    resultado = page.evaluate(JS_LIMPIAR_SLICER, aria_label)
    if resultado == 'ok':
        time.sleep(0.8)
        print(f"  Slicer '{aria_label}' limpiado con borrador")
        return True

    # Fallback: abrir el slicer y doble-clic en "Seleccionar todo"
    print(f"  Borrador no encontrado para '{aria_label}' ({resultado}), usando Seleccionar todo")
    if not abrir_slicer(page, aria_label):
        return False

    JS_CLICK_SELECT_ALL = f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            const container = body?.closest('.slicerContainer');
            const selectAll = container?.querySelector('.slicerItemContainer[data-row-index="0"]');
            if (!selectAll) return false;
            selectAll.click();
            return true;
        }}
    """
    ok1 = page.evaluate(JS_CLICK_SELECT_ALL)
    time.sleep(1.0)
    n_sel = page.evaluate(f"""
        () => {{
            const target = "{aria_label}".normalize('NFC').toLowerCase();
            const body = [...document.querySelectorAll('.slicerBody')].find(
                e => (e.getAttribute('aria-label') || '').normalize('NFC').toLowerCase() === target
            );
            const container = body?.closest('.slicerContainer');
            return container?.querySelectorAll('.slicerItemContainer[aria-selected="true"]').length ?? -1;
        }}
    """)
    if n_sel != 0:
        page.evaluate(JS_CLICK_SELECT_ALL)
        time.sleep(1.0)
    page.keyboard.press("Escape")
    time.sleep(0.5)
    return bool(ok1)


# =============================================================================
# NRC: verificar existencia de NRCs para cursos no proyectados
# =============================================================================

JS_TIENE_NRC = """
() => {
    // Misma estructura que GENERAL: buscar celdas con contenido en el cuerpo de la tabla
    const bodyContainer = document.querySelector('.bodyCells');
    if (bodyContainer) {
        const cells = bodyContainer.querySelectorAll('.cell, .pivotTableCellWrap');
        return [...cells].some(el => el.textContent.trim().length > 0);
    }
    // Fallback: ARIA gridcells
    const gridcells = document.querySelectorAll('[role="gridcell"]');
    return [...gridcells].some(el => el.textContent.trim().length > 0);
}
"""


def verificar_nrc_existe(page) -> bool:
    try:
        return bool(page.evaluate(JS_TIENE_NRC))
    except Exception as e:
        print(f"  Error verificando NRC: {e}")
        return False


def buscar_nrc_cursos(page, cursos_sin_svaproy: list[dict]) -> bool:
    """
    Navega a 'Reporte de NRCs u Horario', verifica PERIODO una vez (sin re-clic si ya
    está marcado) y por cada curso limpia MATERIA/NÚMERO con el borrador antes de filtrar.
    Retorna True en cuanto encuentra un NRC (fail fast).
    """
    print("\n--- REPORTE DE NRCs ---")
    navegar_seccion(page, "NRCs")

    # PERIODO: solo hace clic si no está ya seleccionado
    if not verificar_slicer_seleccionado(page, "PERIODO", BANNER_PERIODO):
        print(f"  No se pudo confirmar PERIODO {BANNER_PERIODO}")
        return False
    time.sleep(2)   # espera breve solo para estabilidad, no recarga completa

    for item in cursos_sin_svaproy:
        codigo = item.get("codigo", "")
        nombre = item.get("nombre", "")
        if not codigo or len(codigo) < 6:
            print(f"  {nombre}: código inválido '{codigo}', omitido")
            continue

        # "INVE01010" → materia="INVE", numero="01010"
        materia = codigo[:-5].strip()
        numero  = codigo[-5:]
        print(f"  NRC: {codigo} ({nombre}) | materia={materia} numero={numero}")

        # Limpiar con borrador y seleccionar MATERIA
        limpiar_slicer_borrador(page, "MATERIA")
        if not seleccionar_en_slicer(page, "MATERIA", materia):
            print(f"  No se pudo seleccionar MATERIA '{materia}'")
            continue

        # Limpiar con borrador y seleccionar NÚMERO DE CURSO
        limpiar_slicer_borrador(page, "NÚMERO DE CURSO")
        if not seleccionar_en_slicer(page, "NÚMERO DE CURSO", numero):
            print(f"  No se pudo seleccionar NÚMERO DE CURSO '{numero}'")
            continue

        print(f"  Esperando {WAIT_NRC}s para carga NRC...")
        time.sleep(WAIT_NRC)

        if verificar_nrc_existe(page):
            print(f"  *** NRC ENCONTRADO: {codigo} → ERROR DEL SISTEMA ***")
            return True
        print(f"  Sin NRC para {codigo}")

    return False


EXCEL_TOTAL   = "analisis_total.xlsx"
EXCEL_REVISAR = "para_revisar.xlsx"


def guardar_para_revisar(filas: list[dict]):
    """Guarda cursos con prerrequisitos no evaluables. Cada dict: {id, codigo, curso, prerequisitos}"""
    if not filas:
        return
    path = Path(EXCEL_REVISAR)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revisar"
        ws.append(["ID", "Código", "Curso", "Prerrequisitos"])

    for f in filas:
        ws.append([f["id"], f.get("codigo", ""), f["curso"], f["prerequisitos"]])
    try:
        wb.save(path)
        print(f"  para_revisar.xlsx: {len(filas)} curso(s) agregado(s)")
    except PermissionError:
        alt = path.with_stem(path.stem + "_" + datetime.now().strftime("%H%M%S"))
        wb.save(alt)
        print(f"  ADVERTENCIA: archivo abierto, guardado como {alt.name}")


def guardar_analisis_total(student_id: str, error_sistema: bool):
    path = Path(EXCEL_TOTAL)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Total"
        ws.append(["ID", "ERROR_DEL_SISTEMA"])

    ws.append([student_id, "Sí" if error_sistema else "No"])
    try:
        wb.save(path)
        print(f"  analisis_total.xlsx: {student_id} → {'ERROR' if error_sistema else 'OK'}")
    except PermissionError:
        alt = path.with_stem(path.stem + "_" + datetime.now().strftime("%H%M%S"))
        wb.save(alt)
        print(f"  ADVERTENCIA: archivo abierto, guardado como {alt.name}")


# =============================================================================
# Flujo principal
# =============================================================================

def run():
    if TEST_MODE and TEST_IDS:
        ids = TEST_IDS
        print(f"MODO PRUEBA - IDs forzados: {ids}")
    else:
        ids = leer_ids(INPUT_FILE)
        if TEST_MODE:
            ids = ids[:1]
            print(f"MODO PRUEBA - primer ID del CSV: {ids[0]}")
        else:
            print(f"IDs a procesar: {len(ids)}")

    with open(LOG_FILE, "w", encoding="utf-8") as log_file, sync_playwright() as p:
        try:
            browser = p.chromium.launch(channel="chrome", headless=False, slow_mo=80)
            print("  Navegador: Chrome")
        except Exception:
            try:
                browser = p.chromium.launch(channel="msedge", headless=False, slow_mo=80)
                print("  Navegador: Edge (Chrome no encontrado)")
            except Exception:
                browser = p.chromium.launch(headless=False, slow_mo=80)
                print("  Navegador: Chromium genérico")

        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        print("\nAbriendo Power BI...")
        page.goto(POWER_BI_URL, wait_until="domcontentloaded", timeout=60000)
        input("\n>>> Inicia sesion y presiona ENTER <<<\n")
        time.sleep(8)

        # Abrir Banner y navegar a SVAPROY
        banner_page = context.new_page()
        print("\nAbriendo Banner en pestaña nueva...")
        banner_page.goto(BANNER_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)
        print(f"  Banner URL: {banner_page.url} | Titulo: {banner_page.title()}")

        print("\nNavegando a SVAPROY...")
        if not navegar_svaproy(banner_page):
            print("  ADVERTENCIA: no se pudo abrir SVAPROY, revisa manualmente.")
            input("\n>>> Revisa Banner y presiona ENTER para continuar <<<\n")

        # Volver foco a Power BI
        page.bring_to_front()

        # Ir a GENERAL al inicio
        navegar_seccion(page, "GENERAL")

        for i, student_id in enumerate(ids, 1):
            print(f"\n{'='*60}")
            print(f"[{i}/{len(ids)}] ALUMNO: {student_id}")
            print('='*60)

            programa_code = None

            # ------------------------------------------------------------------
            # PASO 1: GENERAL — seleccionar ID y extraer datos
            # ------------------------------------------------------------------
            print("\n--- GENERAL ---")
            if seleccionar_en_slicer(page, "ID", student_id):
                # Verificar checkbox
                item = page.locator(f'.slicerItemContainer[title="{student_id}"]').first
                try:
                    item.wait_for(state="visible", timeout=5000)
                    if item.get_attribute("aria-selected") != "true":
                        item.click()
                        time.sleep(1)
                except Exception:
                    pass

                print(f"  Esperando {WAIT_CARGA}s para carga...")
                time.sleep(WAIT_CARGA)

                data = extraer_datos_general(page, student_id)
                if data:
                    guardar_general(student_id, data)
                    programa_code = obtener_programa_code(data)
                    print(f"  Programa code: {programa_code}")
                    log(f"[{i}] GENERAL OK: {student_id} | programa: {programa_code}", log_file)
                else:
                    log(f"[{i}] GENERAL SIN DATOS: {student_id}", log_file)
            else:
                log(f"[{i}] ID NO ENCONTRADO EN SLICER: {student_id}", log_file)
                continue

            # ------------------------------------------------------------------
            # PASO 2: DETALLE DEL ÁREA — confirmar ID y buscar cursos N
            # ------------------------------------------------------------------
            print("\n--- DETALLE DEL AREA ---")
            navegar_seccion(page, "DETALLE DEL")

            # Confirmar que el ID sigue seleccionado (o re-seleccionarlo)
            if not verificar_slicer_seleccionado(page, "ID", student_id):
                seleccionar_en_slicer(page, "ID", student_id)

            print(f"  Esperando {WAIT_CARGA}s para carga...")
            time.sleep(WAIT_CARGA)

            cursos_n = extraer_cursos_N_con_scroll(page)

            if not cursos_n:
                log(f"[{i}] SIN CURSOS N: {student_id}", log_file)
                navegar_seccion(page, "GENERAL")
                continue

            log(f"[{i}] CURSOS N ({len(cursos_n)}): {cursos_n}", log_file)

            # ------------------------------------------------------------------
            # PASO 3: CURSOS PRINCIPALES — buscar prerrequisitos
            # ------------------------------------------------------------------
            if not programa_code:
                log(f"[{i}] SIN CODIGO DE PROGRAMA, omitiendo CURSOS PRINCIPALES", log_file)
                navegar_seccion(page, "GENERAL")
                continue

            print("\n--- CURSOS PRINCIPALES ---")
            # "CURSOS PRINCIPALES" está anidado bajo "Reporte de Planes de Estudio"
            navegar_seccion(page, "Reporte de Planes de Estudio")
            navegar_seccion(page, "CURSOS PRINCIPALES")
            # Esperar que el slicer esté listo (sin sleep fijo)
            try:
                page.wait_for_selector('.slicer-dropdown-menu', timeout=20000, state='visible')
            except PWTimeout:
                print("  Timeout esperando slicer, reintentando en 5s...")
                time.sleep(5)

            # Limpiar slicer y seleccionar solo el programa del alumno
            if not seleccionar_programa_unico(page, "Programa Académico", programa_code):
                log(f"[{i}] NO SE PUDO FILTRAR PROGRAMA {programa_code}", log_file)
                navegar_seccion(page, "GENERAL")
                continue

            # Esperar que la tabla cargue con los datos del nuevo programa
            print(f"  Esperando {WAIT_CARGA}s para carga del programa...")
            time.sleep(WAIT_CARGA)

            mapa_prereqs, mapa_codigos = extraer_mapa_prereqs(page)
            # Normalizar claves a mayúsculas para comparar con los N-cursos de DETALLE
            mapa_upper = {k.upper(): v for k, v in mapa_prereqs.items()}
            cursos_n_set = {c.upper() for c in cursos_n}
            # Reverse map: nombre (mayúsculas) → código (ej. "RAZONAMIENTO...": "INVE01010")
            nombre_upper_to_codigo = {v.upper(): k for k, v in mapa_codigos.items()}
            print(f"  Muestra claves mapa: {list(mapa_upper.keys())[:4]}")

            filas_analisis = []
            for curso in cursos_n:
                if es_curso_excluido(curso):
                    print(f"  {curso} -> [EXCLUIDO electivo/libre/minor]")
                    continue
                prereqs = mapa_upper.get(curso.upper(), '')
                proyectar = calcular_proyectar(prereqs, mapa_codigos, cursos_n_set)
                codigo_curso = nombre_upper_to_codigo.get(curso.upper(), '')
                print(f"  {curso} [{codigo_curso}] -> [{proyectar}] prereqs: {prereqs or '(ninguno)'}")
                filas_analisis.append({
                    "id": student_id,
                    "codigo": codigo_curso,
                    "curso": curso,
                    "prerequisitos": prereqs,
                    "proyectar": proyectar,
                })

            guardar_analisis(filas_analisis)
            log(f"[{i}] ANALISIS GUARDADO: {student_id} | {len(filas_analisis)} cursos", log_file)

            filas_revisar = [f for f in filas_analisis if f.get("proyectar") == "Revisar"]
            if filas_revisar:
                guardar_para_revisar(filas_revisar)
                log(f"[{i}] PARA REVISAR: {len(filas_revisar)} curso(s) con prerreqs no evaluables", log_file)

            # ------------------------------------------------------------------
            # PASO 4: SVAPROY — llenar formulario con ID, periodo y plan
            # ------------------------------------------------------------------
            cursos_proyectar = [
                {"nombre": f["curso"], "codigo": f["codigo"]}
                for f in filas_analisis if f["proyectar"] == "Para proyectar"
            ]
            cursos_sin_svaproy = []

            # ------------------------------------------------------------------
            # PASO 4: SVAPROY — llenar formulario con ID, periodo y plan
            # ------------------------------------------------------------------
            if cursos_proyectar:
                print(f"\n--- SVAPROY ({len(cursos_proyectar)} cursos a proyectar) ---")
                banner_page.bring_to_front()
                ok_svaproy = rellenar_svaproy(banner_page, student_id, BANNER_PERIODO)
                if ok_svaproy:
                    filas_svaproy = exportar_svaproy(banner_page)
                    log(f"[{i}] SVAPROY LEIDO: {len(filas_svaproy)} filas", log_file)
                    cursos_sin_svaproy = guardar_analisis_svaproy(student_id, cursos_proyectar, filas_svaproy)
                else:
                    cursos_sin_svaproy = cursos_proyectar
                log(f"[{i}] SVAPROY {'OK' if ok_svaproy else 'ERROR'}: {student_id}", log_file)
                page.bring_to_front()
            else:
                log(f"[{i}] SVAPROY omitido: sin cursos Para proyectar", log_file)

            # ------------------------------------------------------------------
            # PASO 5: NRC — verificar si cursos sin SVAPROY tienen NRC disponible
            # ------------------------------------------------------------------
            if cursos_sin_svaproy:
                error_sistema = buscar_nrc_cursos(page, cursos_sin_svaproy)
                log(f"[{i}] NRC revisados: {len(cursos_sin_svaproy)} cursos", log_file)
            else:
                error_sistema = False
                log(f"[{i}] NRC omitido: sin cursos fuera de SVAPROY", log_file)

            guardar_analisis_total(student_id, error_sistema)
            log(f"[{i}] ERROR_SISTEMA: {'Sí' if error_sistema else 'No'} | {student_id}", log_file)

            # Volver a GENERAL para el siguiente alumno
            navegar_seccion(page, "GENERAL")

        # Resumen final
        log(f"\n{'='*50}\nBot finalizado. Procesados: {len(ids)}", log_file)
        input("\nPresiona ENTER para cerrar el navegador.")
        browser.close()


if __name__ == "__main__":
    run()
