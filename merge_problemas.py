"""
Cruza analisis_total.xlsx con dos columnas de la encuesta de errores:
  - "¿Tuviste algún problema durante la simulación?" (selección múltiple)
  - "Comentario breve (opcional):"

Uso:
  1. Descarga el Excel de SharePoint y guárdalo como 'encuesta_errores.xlsx'
     en la misma carpeta que este script (o ajusta ENCUESTA_FILE abajo).
  2. python merge_problemas.py

Salida: analisis_total_con_problemas.xlsx
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl import Workbook

# ── Configuración ─────────────────────────────────────────────────────────────

ANALISIS_FILE  = "analisis_total.xlsx"
ENCUESTA_FILE  = "encuesta_errores.xlsx"
OUTPUT_FILE    = "analisis_total_con_problemas.xlsx"

# Fragmentos de texto que identifican cada columna de la encuesta (case-insensitive).
# Ajusta si el encabezado exacto es distinto.
FRAG_PROBLEMA   = "tuviste"          # columna "¿Tuviste algún problema..."
FRAG_COMENTARIO = "comentario"       # columna "Comentario breve (opcional):"

# Columna de ID en la encuesta (None = detectar automáticamente).
ENCUESTA_ID_COL = None

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalizar_id(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def buscar_col(headers: list[str], fragmento: str) -> int | None:
    """Devuelve el índice de la primera columna cuyo header contiene el fragmento."""
    frag = fragmento.lower()
    for i, h in enumerate(headers):
        if frag in h.lower():
            return i
    return None


def leer_encuesta(path: str):
    """
    Devuelve:
      mapa  : {id_alumno: {"problema": str, "comentario": str}}
      labels: (label_problema, label_comentario)  — headers reales para las columnas de salida
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[1]]
    if not headers:
        print(f"  ERROR: {path} no tiene encabezados en la fila 1.")
        sys.exit(1)

    print(f"  Encabezados encuesta ({len(headers)}): {headers}")

    # Columna ID
    if ENCUESTA_ID_COL is not None:
        id_idx = ENCUESTA_ID_COL - 1
    else:
        # Prioridad: columna exactamente "DNI" → luego parciales
        exacto = [i for i, h in enumerate(headers) if h.strip().upper() == "DNI"]
        if exacto:
            candidatos = exacto
        else:
            candidatos = [i for i, h in enumerate(headers)
                          if any(p in h.lower() for p in ("dni", "alumno", "codigo", "código"))]
        id_idx = candidatos[0] if candidatos else 0
    print(f"  Columna ID: [{id_idx+1}] '{headers[id_idx]}'")

    # Columna problema
    prob_idx = buscar_col(headers, FRAG_PROBLEMA)
    if prob_idx is None:
        print(f"  ADVERTENCIA: no se encontró columna con '{FRAG_PROBLEMA}'. Quedará vacía.")
    else:
        print(f"  Columna problema: [{prob_idx+1}] '{headers[prob_idx]}'")

    # Columna comentario
    com_idx = buscar_col(headers, FRAG_COMENTARIO)
    if com_idx is None:
        print(f"  ADVERTENCIA: no se encontró columna con '{FRAG_COMENTARIO}'. Quedará vacía.")
    else:
        print(f"  Columna comentario: [{com_idx+1}] '{headers[com_idx]}'")

    mapa = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_val = normalizar_id(row[id_idx] if id_idx < len(row) else None)
        if not id_val:
            continue

        prob = ""
        if prob_idx is not None and prob_idx < len(row) and row[prob_idx] is not None:
            prob = str(row[prob_idx]).strip()

        com = ""
        if com_idx is not None and com_idx < len(row) and row[com_idx] is not None:
            com = str(row[com_idx]).strip()

        mapa[id_val] = {"problema": prob, "comentario": com}

    label_prob = headers[prob_idx] if prob_idx is not None else "Problema"
    label_com  = headers[com_idx]  if com_idx  is not None else "Comentario"

    print(f"  Registros en encuesta: {len(mapa)}")
    return mapa, label_prob, label_com


def leer_analisis(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    filas = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    print(f"  Filas en analisis_total: {len(filas)}")
    return headers, filas


def detectar_col_id(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if h.upper() in ("ID", "ID ALUMNO", "ALUMNO"):
            return i
    return 0


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    print("=" * 65)
    print("  MERGE: analisis_total + encuesta de errores")
    print("=" * 65)

    for f in (ANALISIS_FILE, ENCUESTA_FILE):
        if not Path(f).exists():
            print(f"\n  ERROR: '{f}' no encontrado en {Path('.').resolve()}")
            if f == ENCUESTA_FILE:
                print(f"  → Descarga el Excel de SharePoint y guárdalo como '{ENCUESTA_FILE}'")
                print(f"     (o ajusta ENCUESTA_FILE al inicio del script)")
            sys.exit(1)

    print(f"\nLeyendo encuesta: {ENCUESTA_FILE}")
    mapa, label_prob, label_com = leer_encuesta(ENCUESTA_FILE)

    print(f"\nLeyendo analisis_total: {ANALISIS_FILE}")
    headers_a, filas_a = leer_analisis(ANALISIS_FILE)
    id_idx_a = detectar_col_id(headers_a)
    print(f"  Columna ID: [{id_idx_a+1}] '{headers_a[id_idx_a]}'")

    print(f"\nCruzando datos...")
    encontrados = 0
    no_encontrados = []

    filas_out = []
    for fila in filas_a:
        id_val = normalizar_id(fila[id_idx_a] if id_idx_a < len(fila) else None)
        datos = mapa.get(id_val, {})
        prob = datos.get("problema", "")
        com  = datos.get("comentario", "")
        if prob or com:
            encontrados += 1
        else:
            no_encontrados.append(id_val)
        filas_out.append(list(fila) + [prob, com])

    print(f"  IDs con datos de encuesta : {encontrados}")
    print(f"  IDs sin datos de encuesta : {len(no_encontrados)}")
    if no_encontrados:
        print(f"  Sin datos: {no_encontrados[:20]}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "analisis_total"

    ws_out.append(list(headers_a) + [label_prob, label_com])
    for fila in filas_out:
        ws_out.append(fila)

    for col in ws_out.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    wb_out.save(OUTPUT_FILE)
    print(f"\nGuardado: {OUTPUT_FILE}  ({len(filas_out)} filas)")
    print("=" * 65)


if __name__ == "__main__":
    run()
