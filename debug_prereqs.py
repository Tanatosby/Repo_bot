"""
Debug de lógica de proyección para cadenas de prerrequisitos específicas.
Lee analisis.xlsx para reconstruir cursos_n_set y mapa_codigos parcial,
luego traza paso a paso por qué cada cadena resulta en 'Para proyectar'.

Uso: python debug_prereqs.py
"""

import re
import openpyxl
from pathlib import Path

STUDENT_ID     = "60167392"
EXCEL_ANALISIS = "analisis.xlsx"

# Cadenas a analizar (copiar de analisis.xlsx o del log)
PREREQS_A_ANALIZAR = [
    "(MATE03003 [11] AND ESTA03001 [11]) OR (ESTA03001 [11] AND MATE03011 [11])",
    "MKTG03021 [11] OR (MKTG03003 [11] AND FINZ03003 [11]) OR NINT03010 [11]",
    "(CCDA03002 [11] AND AUDI03003 [0]) OR (FINZ03003 [11] AND MKTG03004 [11]) OR NINT03011 [11]",
]

# =============================================================================
# Reconstruir cursos_n_set y mapa_codigos desde analisis.xlsx
# =============================================================================

def cargar_datos_alumno(excel_path: str, student_id: str):
    """
    Lee analisis.xlsx y devuelve:
      - cursos_n_set   : set de nombres de cursos con N (en mayúsculas)
      - mapa_codigos   : {codigo: nombre} solo para cursos N del alumno
      - filas          : lista de dicts con toda la info del alumno
    """
    path = Path(excel_path)
    if not path.exists():
        print(f"  ADVERTENCIA: {excel_path} no encontrado")
        return set(), {}, []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(c.value or '').strip() for c in ws[1]]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: (str(v or '').strip()) for i, v in enumerate(row)}
        if d.get('ID', '') == student_id:
            filas.append(d)

    cursos_n_set = {f['Curso no cumple'].upper() for f in filas if f.get('Curso no cumple')}
    # mapa_codigos parcial: solo los cursos N del alumno
    mapa_codigos = {}
    for f in filas:
        codigo = f.get('Código', '').strip()
        nombre = f.get('Curso no cumple', '').strip()
        if codigo and nombre:
            mapa_codigos[codigo] = nombre

    return cursos_n_set, mapa_codigos, filas


# =============================================================================
# Versión verbose de calcular_proyectar
# =============================================================================

def calcular_proyectar_verbose(prereq_str: str, mapa_codigos: dict, cursos_n_set: set):
    """Igual que calcular_proyectar pero imprime cada paso."""

    print(f"\n  Prerrequisito original:")
    print(f"    {prereq_str}")

    if not prereq_str or not prereq_str.strip():
        print(f"  → Sin prerrequisitos → Para proyectar")
        return 'Para proyectar'

    # ── Paso 1: reemplazar cada código por True/False ──────────────────────
    reemplazos = []

    def reemplazar(match):
        codigo = match.group(1)
        nombre = mapa_codigos.get(codigo, '').upper()

        if not nombre:
            # Código no encontrado en mapa → no se puede asumir aprobado → False
            reemplazos.append((codigo, '[NO EN MAPA → False]', 'False'))
            return 'False'
        elif nombre in cursos_n_set:
            # Está entre los cursos N → prerreq NO cumplido → False
            reemplazos.append((codigo, nombre, 'False'))
            return 'False'
        else:
            # Está en el mapa pero NO tiene N → ya aprobado → True
            reemplazos.append((codigo, nombre, 'True'))
            return 'True'

    expr = re.sub(r'([A-Z]{2,5}\d{5})(?:\s*\[\d+\])?', reemplazar, prereq_str)

    print(f"\n  Paso 1 — reemplazo código → True/False:")
    for codigo, nombre, valor in reemplazos:
        flag = "  ← NO EN MAPA" if "NO EN MAPA" in nombre else ""
        print(f"    {codigo:15s} → '{nombre}'{flag}")
        print(f"    {'':15s}    ¿en cursos_n? {'SÍ → False' if valor == 'False' else 'NO → True'}")

    # ── Paso 2: convertir AND/OR ───────────────────────────────────────────
    expr_python = expr.replace(' AND ', ' and ').replace(' OR ', ' or ')

    print(f"\n  Paso 2 — expresión Python:")
    print(f"    {expr_python}")

    # ── Paso 3: eval ──────────────────────────────────────────────────────
    try:
        cumple = eval(expr_python)
        resultado = 'Para proyectar' if cumple else 'No proyectar'
        print(f"\n  Paso 3 — eval: {cumple} → {resultado}")

        # ── Diagnóstico ───────────────────────────────────────────────────
        codigos_no_en_mapa = [c for c, n, v in reemplazos if "NO EN MAPA" in n]
        if codigos_no_en_mapa:
            print(f"\n  Códigos no encontrados en el plan → False:")
            for c in codigos_no_en_mapa:
                print(f"        {c}")

        return resultado

    except Exception as e:
        print(f"\n  Paso 3 — eval FALLÓ: {e} → Revisar")
        return 'Revisar'


# =============================================================================
# Main
# =============================================================================

def run():
    print(f"{'='*65}")
    print(f"  DEBUG PRERREQUISITOS — Alumno: {STUDENT_ID}")
    print(f"{'='*65}")

    cursos_n_set, mapa_codigos, filas = cargar_datos_alumno(EXCEL_ANALISIS, STUDENT_ID)

    print(f"\n  Datos cargados de {EXCEL_ANALISIS}:")
    print(f"    Filas del alumno   : {len(filas)}")
    print(f"    Cursos N (set)     : {len(cursos_n_set)} cursos")
    print(f"    Mapa códigos       : {len(mapa_codigos)} entradas (solo cursos N del alumno)")
    print(f"\n  Cursos con N ({len(cursos_n_set)}):")
    for nombre in sorted(cursos_n_set):
        # Buscar código asociado
        codigo = next((k for k, v in mapa_codigos.items() if v.upper() == nombre), '?')
        print(f"    {codigo:15s} {nombre}")

    print(f"\n{'='*65}")
    print(f"  ANÁLISIS DE CADENAS")
    print(f"{'='*65}")

    for i, prereq in enumerate(PREREQS_A_ANALIZAR, 1):
        print(f"\n--- Cadena {i} ---")
        calcular_proyectar_verbose(prereq, mapa_codigos, cursos_n_set)

    print(f"\n{'='*65}")
    print(f"  NOTA: 'NO EN MAPA' → el código no aparece en CURSOS PRINCIPALES")
    print(f"  del plan del alumno. Se trata como False (no proyectar).")
    print(f"  Si el curso ya estaba aprobado pero no aparece en el mapa,")
    print(f"  el problema es que CURSOS PRINCIPALES no lo listó.")
    print(f"{'='*65}")


if __name__ == "__main__":
    run()
